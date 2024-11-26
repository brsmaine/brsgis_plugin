import openpyxl
import psycopg2


def readFile(file, geocode):
    if str(geocode) == '15010':
        file = file + "AlnaExport.xlsx"
    elif str(geocode) == '23020':
        file = file + "BathExport.xlsx"
    elif str(geocode) == '15020':
        file = file + "Boothbay_Export.xlsx"
    elif str(geocode) == '15030':
        file = file + "BBHExport.xlsx"
    elif str(geocode) == '15040':
        file = file + 'BremenExport.xlsx'
    elif str(geocode) == '15050':
        file = file + "BristolExport.xlsx"
    elif str(geocode) == '15060':
        file = file + "DamariscottaExport.xlsx"
    elif str(geocode) == '15070':
        file = file + "DresdenExport.xlsx"
    elif str(geocode) == '15080':
        file = file + "EdgecombExport.xlsx"
    elif str(geocode) == '23050':
        file = file + "GeorgetownExport.xlsx"
    elif str(geocode) == '15110':
        file = file + "NewcastleExport.xlsx"
    elif str(geocode) == '15120':
        file = file + "NobleboroExport.xlsx"
    elif str(geocode) == '23060':
        file = file + "PhippsburgExport.xlsx"
    elif str(geocode) == '23070':
        file = file + "RichmondExport.xlsx"
    elif str(geocode) == '23090':
        file = file + "WestBathExport.xlsx"
    elif str(geocode) == '15150':
        file = file + "SouthportExport.xlsx"
    elif str(geocode) == '23080':
        file = file + "TopshamExport.xlsx"
    elif str(geocode) == '15160':
        file = file + "WaldoboroExport.xlsx"
    elif str(geocode) == '15170':
        file = file + "WestportExport.xlsx"
    elif str(geocode) == '15180':
        file = file + "WhitefieldExport.xlsx"
    elif str(geocode) == '15190':
        file = file + "WiscassetExport.xlsx"
    elif str(geocode) == '23100':
        file = file + "WoolwichExport.xlsx"
    elif str(geocode) == '05110':
        file = file + "HarpswellExport.xlsx"
    else:
        pass
    wb_obj = openpyxl.load_workbook(file)
    sheet_obj = wb_obj.active
    # cell_obj = sheet_obj.cell(row=1, column=1)

    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    taxData = []
    allData = []
    error_count = 0

    for j in range(2, max_row + 1):

        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=j, column=i)
            taxData.append(str(cell_obj.value))
        allData.append(taxData)
        taxData = []

    for taxBill in allData:

        map_bk_lot = str(taxBill[2])
        if map_bk_lot == '':
            pass
        else:
            locusAddress = taxBill[3]
            name = taxBill[4]
            mailingAddress = taxBill[5]
            address = mailingAddress.replace("\n", " | ")
            deeds = taxBill[7]
            pages = deeds.split(" ")
            numDeeds = len(pages)
            sPages = str(pages)

            for x in range(0, numDeeds):
                if x == 0:
                    bookpage = str(pages[x])
                elif x == 1:
                    bookpage = bookpage + ',' + str(pages[x])

            if str(geocode) == '15010':
                addressName = 'N/A'
                addressStreet = 'N/A'
                addressStreet2 = 'N/A'
                addressCity = 'N/A'
                addressState = 'N/A'
                addressZip = 'N/A'
                # print('DEED TO KEEP: ' + bookpage + ' | ALL DEED INFO: ' + str(pages))
                # print(locusAddress)
                print('.')
                updateTable(map_bk_lot, geocode, name, locusAddress, address, mailingAddress, bookpage, deeds,
                            addressName, addressStreet, addressStreet2, addressCity, addressState, addressZip)

            else:
                addressName = taxBill[11]
                addressStreet = taxBill[12]
                addressStreet2 = taxBill[13]
                addressCity = taxBill[14]
                addressState = taxBill[15]
                addressZip = taxBill[16]

                # print('DEED TO KEEP: ' + bookpage + ' | ALL DEED INFO: ' + str(pages))
                # print(bookpage)
                print('.')
                updateTable(map_bk_lot, geocode, name, locusAddress, address, mailingAddress, bookpage, deeds,
                            addressName, addressStreet, addressStreet2, addressCity, addressState, addressZip)


def updateTable(map_bk_lot, geocode, owner, locusAddress, address, mailingAddress, bookpage, deeds, addressee, street1,
                street2, city, state, zip):
    try:
        connection = psycopg2.connect(user="postgres",
                                      password="BRSMAinE4445",
                                      host="192.168.1.101",
                                      port="5432",
                                      database="BRS_GIS_PRD")

        cursor = connection.cursor()
        if str(geocode) == '15010':  # Alna uses a unique template
            sql = """Update parcels_aux set locusaddress = %s,
                        owner1 = %s,
                        bookpage = %s,
                        rawdeeds = %s
                        where map_bk_lot = %s and geocode = %s"""
            cursor.execute(sql, (locusAddress, owner, bookpage, deeds, map_bk_lot, geocode))
            connection.commit()

        else:  # All other towns use a common template

            sql = """Update parcels_aux set locusaddress = %s,
                        owner1 = %s,
                        mailingaddress = %s,
                        bookpage = %s,
                        rawdeeds = %s,
                        addressee = %s,
                        formattedaddress = %s,
                        own_street1 = %s,
                        own_street2 = %s,
                        own_city = %s,
                        own_state = %s,
                        own_zip = %s
                        where map_bk_lot = %s and geocode = %s"""
            cursor.execute(sql, (locusAddress, owner, mailingAddress, bookpage, deeds, addressee, address, street1,
                                 street2, city, state, zip, map_bk_lot, geocode))
            connection.commit()

    except (Exception, psycopg2.Error) as error:
        print("Error in update operation", error)

    finally:
        # closing database connection.
        if (connection):
            cursor.close()
            connection.close()
            # print("PostgreSQL connection is closed")


class townImport():
    # startMain
    path = "D:\\LeightonProjects\\GEO DATA\\Town_Data\\0-TownExports\\Effective\\"
    # readFile(path, '15010')  # Alna, VERIFIED.
    # readFile(path, '23020')  # Bath, VERIFIED.
    # readFile(path, '15020')  # Boothbay, VERIFIED.
    # readFile(path, '15030')  # Boothbay Harbor, VERIFIED.
    # readFile(path, '15050')  # Bristol, VERIFIED.
    # readFile(path, '15060')  # Damariscotta, VERIFIED.
    # readFile(path, '15070')  # Dresden, VERIFIED.
    # readFile(path, '15080')  # Edgecomb, VERIFIED.
    # readFile(path, '23050')  # Georgetown, VERIFIED.
    # readFile(path, '15110')  # Newcastle, VERIFIED.
    # readFile(path, '23060')  # Phippsburg, VERIFIED.
    # readFile(path, '23070')  # Richmond, VERIFIED.
    # readFile(path, '23080')  # Topsham, VERIFIED.
    # readFile(path, '15160')  # Waldoboro, VERIFIED.
    # readFile(path, '15170')  # Westport Island, VERIFIED.
    # readFile(path, '15180')  # Whitefield, VERIFIED.
    # readFile(path, '15150') # Southport, PROCESSED/PENDING.
    # readFile(path, '15120') # Nobleboro, PROCESSED/PENDING - REWORK NEEDED.
    # readFile(path, '15040') # Bremen, PROCESSED/PENDING.
    # readFile(path, '23090')  # West Bath, PROCESSED/PENDING.
    # readFile(path, '15190')  # Wiscasset, PROCESSED/PENDING.
    # readFile(path, '23100')  # Woolwich, PROCESSED/PENDING.
    readFile(path, '05110')  # Harpswell, PROCESSED/PENDING.
