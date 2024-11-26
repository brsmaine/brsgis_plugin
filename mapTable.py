class brsgis_printMapTable(object):

    def __init__(self, iface):
        self.iface = iface

    def initGui(self):

        self.action = QAction("PMT", self.iface.mainWindow())
        self.action.triggered.connect(self.run)
        self.action.trigger()

    def run(self):

        cfg0 = 0
        cfg1 = 1

        layerActive = self.iface.activeLayer()

        try:
            feat = self.iface.activeLayer().selectedFeatures()[0]

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "No Selection",
                                 "Please ensure that you have a parcel selected\nand attempt to "
                                 "generate the output again.\n\n"
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return

        fid = feat.id()
        key = str(feat['map_bk_lot'])
        if key == 'NULL':
            key = str(feat['sid'])
        QgsMessageLog.logMessage('DROP: ' + str(key), 'BRS_GIS', level=Qgis.Info)

        self.iface.setActiveLayer(layerActive)

        self.getRelatedWork(feat, cfg0)

        if layerActive.name() == 'brs_jobs':
            import datetime
            relW = self.updateJobRelated(feat)
            year = datetime.datetime.today().strftime('%Y')

            try:
                attribs = self.iface.activeLayer().selectedFeatures()[0]
            except Exception as e:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                QMessageBox.critical(self.iface.mainWindow(), "NO SELECTION!",
                                     "Please ensure that you have a parcel selected\nand attempt to "
                                     "generate the output again.\n\n"
                                     "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                         exc_tb.tb_lineno) + ' ' + str(e))
                return

            jobNo = attribs["job_no"]
            jobYear = '20' + jobNo[:2]

            if jobYear == year:
                year = year
            else:
                year = jobYear

            path = os.path.join("Z:\\", "BRS-DEV", year, jobNo)
            jipath = os.path.join(path, "Job_Info")
            if not os.path.exists(path):
                os.makedirs(path)
            if not os.path.exists(jipath):
                os.makedirs(jipath)

            from openpyxl import load_workbook

            tFile = 'BRS_templates.xlsx'
            fPath = self.resolve(tFile)

            wb = load_workbook(fPath)

            for s in range(len(wb.sheetnames)):
                if wb.sheetnames[s] == 'maptable':
                    break
            wb.active = s
            sheets = wb.sheetnames
            ws = wb.active

            clientName = attribs["client_name"]
            folderName = attribs["folder_name"]
            addr = attribs["locus_addr"]
            town = attribs["town"]

            try:

                map_bk_lot = attribs["map_bk_lot"]
                ogMap = map_bk_lot
                map_bk_lotO = attribs["map_bk_lot"]
                mbl = map_bk_lot.split('-')
                mbLen = len(mbl)

                # NEED TO HANDLE MORE THAN 3 SECTIONS OF MBL
                if mbLen == 1:
                    map_bk_lot = map_bk_lot
                elif mbLen == 2:
                    map_bk_lot = 'Map ' + mbl[0].lstrip('0') + ', Lot ' + mbl[1].lstrip('0')
                elif mbLen == 3:
                    map_bk_lot = 'Map ' + mbl[0].lstrip('0') + ', Lot ' + mbl[1].lstrip('0') + '-' + mbl[2].lstrip('0')

            except Exception as e:
                map_bk_lot = 'N/A'

            jobType = attribs["job_type"]
            jobSubType = attribs["jobSubtype"]

            if str(jobSubType) == 'NULL':
                jobSubType = ''
            else:
                jobSubType = jobSubType

            revNo = attribs["rev_no"]
            county = attribs["county"]
            state = attribs["state"]
            perimeter = attribs["sPerimeter"]
            area = attribs["area"]
            planbook_page = attribs["planbook_page"]
            referrerJ = attribs['job_no']
            zipCode = attribs['zipcode']

            if str(zipCode) == 'NULL':
                zipCode = ''
            else:
                zipCode = zipCode

            lat_lon = attribs['lat_lon']

            if str(clientName) == 'NULL':
                clientName = 'N/A'
            else:
                clientName = clientName

            if str(county) == 'NULL':
                county = 'N/A'
            else:
                county = county

            try:
                ws['A1'] = folderName
                ws['D2'] = town.upper()
                ws['E2'] = county
                ws['F2'] = state
                ws['A3'] = 'Job#: ' + str(jobNo)
                ws['B3'] = 'Rev#: ' + str(revNo)
                ws['C3'] = 'Type: ' + str(jobType)  # + ' | ' + str(jobSubType)
                ws['D3'] = 'Perimeter: ' + str(perimeter) + ' L.Ft'
                ws['E3'] = 'Area: ' + str(area) + ' Ac.'
                ws['F3'] = 'Centroid: ' + str(lat_lon)
                ws['D4'] = zipCode
                ws['A6'] = map_bk_lot
                ws['B6'] = addr
                ws['E6'] = folderName
                ws['B9'] = str(relW)

                layer3 = QgsProject.instance().mapLayersByName('abutters')[0]
                exp = QgsExpression(u'"referrerj" = \'%s\'' % (jobNo))
                request = QgsFeatureRequest(exp)
                request.setSubsetOfAttributes(['referrerj'], layer3.fields())
                request.setFlags(QgsFeatureRequest.NoGeometry)

                aNo = 0
                startCell = 13
                startCellb = 14
                startCellp = 15
                startCelld = 16

                for f in layer3.getFeatures(request):

                    if str(f['map_bk_lot']) == key:
                        owner1 = str(f['owner1'])
                        if owner1 == 'NULL':
                            owner1 = ''

                        formattedaddress = str(f['formattedaddress'])
                        if formattedaddress == 'NULL':
                            formattedaddress = ''
                        locusaddress = str(f['locusaddress'])
                        if locusaddress == 'NULL':
                            locusaddress = ''

                        ownInfo = 'OWNER: ' + owner1 # + ' | BOOKPAGE: ' + str(f['bookpage'])
                        formattedaddress = 'MAIL: ' + formattedaddress
                        alldeeds = str(f['rawdeeds'])
                        if alldeeds == 'NULL':
                            alldeeds = ''

                        ws['A7'] = locusaddress
                        if len(ownInfo) >= 150:
                            ws.row_dimensions[7].height = 45
                            ws['B10'].alignment = Alignment(horizontal='left', vertical='center', text_rotation=0,
                                                    wrap_text=True, shrink_to_fit=False, indent=0)
                        ws['B7'] = ownInfo
                        if len(formattedaddress) >= 150:
                            ws.row_dimensions[10].height = 45
                            ws['B8'].alignment = Alignment(horizontal='left', vertical='center', text_rotation=0,
                                                    wrap_text=True, shrink_to_fit=False, indent=0)
                        ws['B8'] = formattedaddress
                        if len(alldeeds) >= 150:
                            ws.row_dimensions[10].height = 45
                            ws['B10'].alignment = Alignment(horizontal='left', vertical='center', text_rotation=0,
                                                    wrap_text=True, shrink_to_fit=False, indent=0)
                        ws['A10'] = 'All Deeds:'
                        ws['B10'] = alldeeds

                        pass
                    else:

                        QgsMessageLog.logMessage('abutter found: ' + str(f['map_bk_lot']), 'BRS_GIS', level=Qgis.Info)

                        aNo += 1
                        c1 = 'A' + str(startCell)
                        c2 = 'B' + str(startCell)
                        c3 = 'E' + str(startCell)
                        c4 = 'B' + str(startCellp)
                        c5 = 'B' + str(startCellb)
                        c6 = 'A' + str(startCellb)
                        c7 = 'A' + str(startCelld)
                        c8 = 'B' + str(startCelld)

                        map_bk_lotP = f['map_bk_lot']
                        mbl = map_bk_lotP.split('-')
                        mbLen = len(mbl)

                        if mbLen == 1:
                            map_bk_lotP = map_bk_lotP
                        elif mbLen == 2:
                            map_bk_lotP = 'Map ' + mbl[0].lstrip('0') + ', Lot ' + mbl[1].lstrip('0')
                        elif mbLen == 3:
                            map_bk_lotP = 'Map ' + mbl[0].lstrip('0') + ', Lot ' + mbl[1].lstrip('0') + '-' + mbl[2].lstrip(
                                '0')

                        ws[c1] = str(map_bk_lotP)

                        if str(f['locusaddress']) == 'NULL':
                            ownInfo = ''
                            bookpage = ''
                            formattedaddress = ''
                            locusaddress = ''
                            alldeeds = ''
                            if str(f['bookpage']) == 'NULL':
                                bookpage = ''
                                formattedaddress = ''
                                locusaddress = ''
                                alldeeds = ''
                            else:
                                pass
                        else:
                            owner1 = str(f['owner1'])
                            if owner1 == 'NULL':
                                owner1 = ''

                            ownInfo = 'OWNER: ' + owner1  # + ' | BOOKPAGE: ' + str(f['bookpage'])
                            formattedaddress = 'MAIL: ' + str(f['formattedaddress'])
                            locusaddress = str(f['locusaddress'])
                            alldeeds = str(f['rawdeeds'])
                            # bookpage = str(f['bookpage'])
                        try:
                            if len(ownInfo) > 0:
                                if len(ownInfo) >= 115:
                                    ws.row_dimensions[startCell].height = 30
                                    ws[c2].alignment = Alignment(horizontal='left', vertical='center',
                                                                    text_rotation=0, wrap_text=True,
                                                                 shrink_to_fit=False, indent=0)
                                    ws.merge_cells(start_row=startCell, start_column=2, end_row=startCell, end_column=8)
                                if len(formattedaddress) >= 115:
                                    ws.row_dimensions[startCellb].height = 30
                                    ws[c5].alignment = Alignment(horizontal='left', vertical='center',
                                                                    text_rotation=0, wrap_text=True,
                                                                 shrink_to_fit=False, indent=0)
                                    ws.merge_cells(start_row=startCellb, start_column=2, end_row=startCellb, end_column=8)
                                ws[c2] = ownInfo
                                ws[c7] = 'All Deeds:'
                                if len(alldeeds) >= 115:
                                    ws.row_dimensions[startCelld].height = 30
                                    ws[c5].alignment = Alignment(horizontal='left', vertical='center',
                                                                    text_rotation=0,
                                                                    wrap_text=True, shrink_to_fit=False, indent=0)
                                    ws.merge_cells(start_row=startCelld, start_column=2, end_row=startCelld, end_column=8)
                                ws[c8] = alldeeds

                            else:
                                ws[c2] = ''
                        except Exception as e:
                            ws[c2] = ''
                            exc_type, exc_obj, exc_tb = sys.exc_info()
                            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                            QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION!",
                                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                                     exc_tb.tb_lineno) + ' ' + str(e))
                            return

                        layer3.selectByIds([f.id()])
                        self.iface.setActiveLayer(layer3)
                        sA = self.iface.activeLayer().selectedFeatures()[0]
                        oid = sA['gid']
                        QgsMessageLog.logMessage('ABUTTER SELECTED: ' + str(oid), 'BRS_GIS', level=Qgis.Info)

                        # try:
                        self.getRelatedWork(sA, cfg0)
                        jobNo = str(jobNo)
                        # QgsMessageLog.logMessage('jobNo: ' + str(jobNo), 'BRS_GIS', level=Qgis.Info)
                        relAW = self.updateAbutterRelated(sA, jobNo)

                        QgsMessageLog.logMessage('RELATED AW: ' + str(relAW), 'BRS_GIS', level=Qgis.Info)
                        if len(relAW) >= 115:
                            ws.row_dimensions[startCellp].height = 30
                            ws[c4].alignment = Alignment(horizontal='left', vertical='center',
                                                         text_rotation=0,
                                                         wrap_text=True, shrink_to_fit=False, indent=0)
                        ws[c4] = relAW
                        ws[c5] = formattedaddress
                        ws[c6] = locusaddress
                        ws[c7] = 'All Deeds:'
                        ws[c8] = alldeeds
                        startCell += 4
                        startCellb += 4
                        startCellp += 4
                        startCelld += 4
                        relAW = ''

                        # QgsMessageLog.logMessage('RELATED W: ' + relW, 'BRS_GIS',level=Qgis.Info)

                        # # except Exception as e:
                        #     # QgsMessageLog.logMessage('NO RELATED WORK found: ' + str(f['map_bk_lot']), 'BRS_GIS',level=Qgis.Info)
                        #     # delRow = startCellp + 2
                        #     # ws.delete_rows(delRow,2)
                        #     startCell += 4
                        #     startCellb += 4
                        #     startCellp += 4
                        #     startCelld += 4
                        #     QgsMessageLog.logMessage('NO RELATED WORK.', 'BRS_GIS', level=Qgis.Info)
                        #     exc_type, exc_obj, exc_tb = sys.exc_info()
                        #     fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        #     # QMessageBox.critical(self.iface.mainWindow(), "EXCEPTION",
                        #     #                      "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                        #     #                          exc_tb.tb_lineno) + ' ' + str(e))
                        #     relAW = 'N/A'
                        #     ws[c4] = relAW
                        #     ws[c5] = formattedaddress
                        #     ws[c6] = locusaddress
                        #     ws[c7] = 'All Deeds:'
                        #     ws[c8] = alldeeds
                        #     # rd = ws.row_dimensions[startCellj]
                        #     # rd.height = 1
                        #     pass

            except Exception as e:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                QMessageBox.critical(self.iface.mainWindow(), "ERROR!",
                                     "Please ensure that you have a parcel selected\nand attempt to "
                                     "generate the output again.\n\n"
                                     "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                         exc_tb.tb_lineno) + ' ' + str(e))

            delRow = startCell
            ws.delete_rows(delRow, 200)

            # QgsMessageLog.logMessage('cell/aNo: ' + str(startCell) + '/' + str(aNo), 'BRS_GIS',level=Qgis.Info)

            for s in sheets:

                if s != 'maptable':
                    sheet_name = wb.get_sheet_by_name(s)
                    wb.remove_sheet(sheet_name)

            # Save the file
            mtfile = str(jipath) + "\\" + jobNo + "_MapTable_" + datetime.datetime.today().strftime(
                '%Y.%m.%d') + ".xlsx"
            QgsMessageLog.logMessage('Saving files: ' + mtfile + '...', 'BRS_GIS', level=Qgis.Info)
            try:
                wb.save(mtfile)

            except Exception as e:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                QMessageBox.critical(self.iface.mainWindow(), "File Open",
                                     "Please ensure that you do not have today's mapTable open in Excel\nand attempt to "
                                     "generate the output again.\n\n"
                                     "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(exc_tb.tb_lineno) + ' ' + str(e))
                                                 
            year = datetime.datetime.today().strftime('%Y')
            path = os.path.join("Z:\\", "BRS-DEV", year, jobNo)  # need to programattically grab year
            jipath = os.path.join(path, "Job_Info")
            
            yfile = str(jipath) + "\\" + jobNo + "_YellowSheet_" + datetime.datetime.today().strftime(
                '%Y.%m.%d') + ".xlsx"
            pfile = str(jipath) + "\\" + jobNo + "_YellowSheet_" + datetime.datetime.today().strftime(
                '%Y.%m.%d') + ".pdf"
            pfile2 = str(jipath) + "\\" + jobNo + "_MapTable_" + datetime.datetime.today().strftime(
                '%Y.%m.%d') + ".pdf"    
            outfile = str(jipath) + "\\" + jobNo + "_YellowSheet&MapTable_" + datetime.datetime.today().strftime(
                '%Y.%m.%d') + ".pdf"

                                         
            #convert to PDF
            excel = client.Dispatch("Excel.Application")
            excel.Interactive = False
            excel.Visible = False
            excel.DisplayAlerts = False
            sheets = excel.Workbooks.Open(yfile)
            wks = sheets.Worksheets[0]
            wks.ExportAsFixedFormat(0,pfile)
            sheets.Close(False)
            sheets = None
            excel.Quit()
            excel = None

            excel2 = client.Dispatch("Excel.Application")
            excel2.Interactive = False
            excel2.Visible = False
            excel2.DisplayAlerts = False
            sheets2 = excel2.Workbooks.Open(mtfile)
            wks2 = sheets2.Worksheets[0]
            wks2.ExportAsFixedFormat(0,pfile2)
            wks2 = None
            sheets2.Close(True)
            sheets2 = None
            excel2.Quit()
            excel2 = None
            
            filenames = [pfile, pfile2]
            #merger = PdfMerger()
            merger = PdfWriter()
            for filename in filenames:
                #merger.append(PdfReader(open(filename, 'rb')))
                merger.append(filename)
            merger.write(outfile)
            merger.close()
            merger = None

            QgsMessageLog.logMessage('Saving duplex and flushing temporary files: ' + outfile + '...', 'BRS_GIS', level=Qgis.Info)

            self.vl = QgsProject.instance().mapLayersByName('brs_jobs')[0]
            self.iface.setActiveLayer(self.vl)    
            target_field = 'job_no' 
            condition = jobNo                   
            
            #re-select job_no by attribute
            self.vl.selectByExpression(f"\"{target_field}\"='{condition}'")          
            self.iface.messageBar().clearWidgets()
            try:
                os.remove(yfile)
                os.remove(pfile)
                os.remove(mtfile)
                os.remove(pfile2)
                
            except Exception as e:
                pass
                
            return

        layerRelated = QgsProject.instance().mapLayersByName('relatedwork')[0]
        layerRelated.setSubsetString('id > 1')

        self.vl.dataProvider().forceReload()
        self.iface.mapCanvas().refresh()
        self.iface.setActiveLayer(layerActive)
        layerActive.selectByIds([fid])
        resetLegend(self)

    def resolve(name, basepath=None):
        if not basepath:
            basepath = os.path.dirname(os.path.realpath(__file__))
        else:
            qPath = os.path.dirname(os.path.realpath(__file__)) + '\\' + basepath
            return qPath

    def identAbutters(self):

        aNo = 0

        self.vl = QgsProject.instance().mapLayersByName('brs_jobs')[0]
        self.iface.setActiveLayer(self.vl)

        try:
            attribs = self.iface.activeLayer().selectedFeatures()[0]

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            QMessageBox.critical(self.iface.mainWindow(), "No Selection",
                                 "Please ensure that you have a parcel selected\nand attempt to "
                                 "generate the output again.\n\n"
                                 "Details: " + str(exc_type) + ' ' + str(fname) + ' ' + str(
                                     exc_tb.tb_lineno) + ' ' + str(e))
            return

        jobNo = attribs["job_no"]
        map_bk_lot = attribs["map_bk_lot"]

    def getRelatedWork(self, feature, cfg):
        if cfg == 0:

            layer = self.iface.activeLayer()

            if layer.name() not in ('brs_jobs', 'abutters'):
                return

            layerJobs = QgsProject.instance().mapLayersByName('brs_jobs')[0]
            layerPlans = QgsProject.instance().mapLayersByName('la_plans')[0]
            layerRelated = QgsProject.instance().mapLayersByName('relatedwork')[0]
            layerAbutters = QgsProject.instance().mapLayersByName('abutters')[0]
            layerSupps = QgsProject.instance().mapLayersByName('brs_supplementals')[0]

            self.fb = QgsProcessingFeedback()
            self.context = QgsProcessingContext

            if self.iface.activeLayer().name() == 'brs_jobs':
                jobNo = feature['job_no']
                map_bk_lot = feature['map_bk_lot']
                town = feature['town']
                if str(map_bk_lot) == 'NULL':
                    pass
                else:
                    QgsMessageLog.logMessage('MAPBKLOT: ' + str(map_bk_lot), 'BRS_GIS', level=Qgis.Info)

                    self.iface.actionToggleEditing().trigger()
                    layerRelated.setSubsetString(u'"job_no" = \'%s\'' % (jobNo))
                    listOfIds = [feat.id() for feat in layerRelated.getFeatures()]
                    layerRelated.dataProvider().deleteFeatures(listOfIds)
                    self.iface.activeLayer().commitChanges()
                    self.iface.setActiveLayer(layerJobs)
                    layerRelated.setSubsetString('id > 1')

                    lId = self.iface.activeLayer().id()
                    varI = QgsProcessingFeatureSourceDefinition(str(lId), True)

                    QgsMessageLog.logMessage('lId: ' + str(lId) + ' ' + str(varI), 'BRS_GIS', level=Qgis.Info)

                    try:
                        processing.runAndLoadResults("qgis:intersection",
                                                 {'INPUT': varI,
                                                  'INPUT_FIELDS': ['job_no', 'map_bk_lot'],
                                                  'OUTPUT': 'memory:tmp_related',
                                                  'OVERLAY': layerPlans,
                                                  'OVERLAY_FIELDS': ['plan_no', 'job', 'old_plan']
                                                  }, feedback=self.fb)
                        self.addRelated(0)
                    except Exception as e:
                        pass
                    # getRelated JOBS for JOB feature
                    try:
                        processing.runAndLoadResults("qgis:intersection",
                                                     {'INPUT': varI,
                                                      'INPUT_FIELDS': ['job_no', 'map_bk_lot'],
                                                      'OUTPUT': 'memory:tmp_related',
                                                      'OVERLAY': layerJobs,
                                                      'OVERLAY_FIELDS': ['plan_no', 'job', 'old_plan']
                                                      }, feedback=self.fb)
                        self.addRelated(0)
                    except Exception as e:
                        pass
                    # getRelated SUPPS for JOB feature
                    try:

                        processing.runAndLoadResults("qgis:intersection",
                                                     {'INPUT': varI,
                                                      'INPUT_FIELDS': ['job_no', 'map_bk_lot'],
                                                      'OUTPUT': 'memory:tmp_related',
                                                      'OVERLAY': layerSupps,
                                                      'OVERLAY_FIELDS': ['plan_no', 'job', 'old_plan']
                                                      }, feedback=self.fb)
                        self.addRelated(0)
                    except Exception as e:
                        pass

            elif self.iface.activeLayer().name() == 'abutters':
                lId = self.iface.activeLayer().id()
                varI = QgsProcessingFeatureSourceDefinition(str(lId), True)
                mbl = feature['map_bk_lot']

                self.iface.actionToggleEditing().trigger()

                # NEED TO CHECK POTENTIAL TOWN OVERLAP ISSUES. add TOWN condition when getting RELATED for ABUTTER?
                layerRelated.setSubsetString(u'"map_bk_lot" = \'%s\'' % (mbl))
                listOfIds = [feat.id() for feat in layerRelated.getFeatures()]
                QgsMessageLog.logMessage('LIST OF IDs: ' + str(listOfIds), 'BRS_GIS', level=Qgis.Info)

                layerRelated.dataProvider().deleteFeatures(listOfIds)
                self.iface.activeLayer().commitChanges()
                self.iface.setActiveLayer(layerAbutters)
                layerRelated.setSubsetString('id > 1')

                # getRelated PLANS for ABUTTER feature
                processing.runAndLoadResults("qgis:intersection",
                                             {'INPUT': varI,
                                              'INPUT_FIELDS': ['job_no', 'map_bk_lot', 'town_parcels'],
                                              'OUTPUT': 'memory:tmp_related',
                                              'OVERLAY': layerPlans,
                                              'OVERLAY_FIELDS': ['plan_no', 'job', 'old_plan']
                                              }, feedback=self.fb)

                self.addRelated(0)

                processing.runAndLoadResults("qgis:intersection",
                                             {'INPUT': varI,
                                              'INPUT_FIELDS': ['job_no', 'map_bk_lot', 'town'],
                                              'OUTPUT': 'memory:tmp_related',
                                              'OVERLAY': layerJobs,
                                              'OVERLAY_FIELDS': ['plan_no', 'job', 'old_plan']
                                              }, feedback=self.fb)

                self.addRelated(0)
                processing.runAndLoadResults("qgis:intersection",
                                             {'INPUT': varI,
                                              'INPUT_FIELDS': ['job_no', 'map_bk_lot', 'town'],
                                              'OUTPUT': 'memory:tmp_related',
                                              'OVERLAY': layerSupps,
                                              'OVERLAY_FIELDS': ['plan_no', 'job', 'old_plan']
                                              }, feedback=self.fb)
                self.addRelated(0)
                self.addRelatedAbutterJobs(mbl)

            else:
                pass
        else:
            pass

    def addRelated(self, cfg):
        # try:

        layerTmpRelated = QgsProject.instance().mapLayersByName('Intersection')[0]
        layerTmpRelated.selectAll()
        self.iface.actionCopyFeatures().trigger()
        self.vl = QgsProject.instance().mapLayersByName('relatedwork')[0]
        self.iface.setActiveLayer(self.vl)
        self.iface.actionToggleEditing().trigger()

        self.iface.actionIdentify().trigger()
        self.iface.actionPasteFeatures().trigger()
        self.iface.activeLayer().commitChanges()
        self.iface.messageBar().clearWidgets()

        QgsProject.instance().removeMapLayer(layerTmpRelated.id())
        self.vl.dataProvider().forceReload()
        self.iface.mapCanvas().refresh()
        self.iface.messageBar().clearWidgets()

    def addRelatedAbutterJobs(self, map):
        # try:
        layerJobs = QgsProject.instance().mapLayersByName('brs_jobs')[0]
        self.iface.setActiveLayer(layerJobs)
        layerJobs.setSubsetString(u'"map_bk_lot" = \'%s\'' % (map))
        layerJobs.selectAll()

        QgsMessageLog.logMessage('JOB MAP FILTER: ' + str(map), 'BRS_GIS', level=Qgis.Info)
        #return
        self.iface.actionCopyFeatures().trigger()
        self.vl = QgsProject.instance().mapLayersByName('relatedwork')[0]
        self.iface.setActiveLayer(self.vl)
        self.iface.actionToggleEditing().trigger()
        self.iface.actionIdentify().trigger()
        self.iface.actionPasteFeatures().trigger()
        self.iface.activeLayer().commitChanges()

        layerJobs.setSubsetString("\"supp_type\" = 'X'")

        self.vl.dataProvider().forceReload()
        self.iface.mapCanvas().refresh()

    def showFeatureCount(layers):

        layer = layers[0]
        if layer.type() == QgsMapLayer.VectorLayer:
            root = QgsProject.instance().layerTreeRoot()
            myLayerNode = root.findLayer(layer.id())
            myLayerNode.setCustomProperty("showFeatureCount", True)

    def updateJobRelated(self, feat):

        self.vl = QgsProject.instance().mapLayersByName('brs_jobs')[0]
        self.iface.setActiveLayer(self.vl)
        jobNo = feat['job_no']
        self.town = feat['town']
        layerRelated = QgsProject.instance().mapLayersByName('relatedwork')[0]
        relatedWork = []
        plans = []
        plans2 = []
        ppval = ''
        pval = ''
        pNo = 0

        layerRelated.setSubsetString(u'"job_no" = \'%s\'' % (jobNo))
        exp = QgsExpression(u'"job_no" = \'%s\'' % (jobNo))
        request = QgsFeatureRequest(exp)
        request.setSubsetOfAttributes(['job_no', 'old_plan'], layerRelated.fields())
        request.setFlags(QgsFeatureRequest.NoGeometry)

        for feat in layerRelated.getFeatures(request):

            pFinal = ''
            oldPlan = feat['old_plan']
            pval = oldPlan

            # QgsMessageLog.logMessage('LOOK HERE MOTHERFUCKER!!! | ' + str(pval), 'BRS_GIS', level=Qgis.Info)

            if str(jobNo) in (str(pval)):
                # QgsMessageLog.logMessage('FUCK YEAH YOU DID! | ' + str(pval), 'BRS_GIS', level=Qgis.Info)
                pass
            else:

                if str(ppval) == str(pval):
                    pval = ''

                elif str(pval) == 'NULL':
                    pval = ''
                    # QgsMessageLog.logMessage('NULL pval: ' + str(pval), 'BRS_GIS', level=Qgis.Info)
                else:
                    pval = pval
                    plans.append(pval)

            pNo += 1
            ppval = pval
            QgsMessageLog.logMessage('UJR_plans: ' + str(plans), 'BRS_GIS', level=Qgis.Info)

        feat = self.iface.activeLayer().selectedFeatures()[0]

        if len(plans) >= 1:
            plans.sort(reverse=True)
            pFinal = str(plans)
            pFinal = pFinal.strip('[')
            pFinal = pFinal.strip(']')
            pFinal = pFinal.replace('\'', '')

            self.iface.actionToggleEditing().trigger()
            layerData = self.vl.dataProvider()
            idx3 = layerData.fieldNameIndex('old_plan_no')
            self.vl.changeAttributeValue(feat.id(), idx3, str(pFinal))
            self.vl.updateFields()
            self.iface.activeLayer().commitChanges()

            plans = []
            pval = ''
            ppval = ''

            # QgsMessageLog.logMessage('pFinal: ' + str(pFinal), 'BRS_GIS', level=Qgis.Info)

        else:
            pFinal = ''
        layerRelated.setSubsetString('id > 1')
        # END getAllRelated for selected JOB
        return pFinal

    def updateAbutterRelated(self, feat, jobNo):

        self.vl = QgsProject.instance().mapLayersByName('abutters')[0]
        self.iface.setActiveLayer(self.vl)
        mbl = feat['map_bk_lot']
        layerRelated = QgsProject.instance().mapLayersByName('relatedwork')[0]
        relatedWork = []
        plans = []
        plans2 = []
        ppval = ''
        pval = ''
        pNo = 0

        layerRelated.setSubsetString(u'"map_bk_lot" = \'%s\'' % (mbl))
        exp = QgsExpression(u'"map_bk_lot" = \'%s\'' % (mbl))
        request = QgsFeatureRequest(exp)
        request.setSubsetOfAttributes(['old_plan'], layerRelated.fields())
        request.setFlags(QgsFeatureRequest.NoGeometry)

        for feat in layerRelated.getFeatures(request):

            oldPlan = feat['old_plan']
            fid = feat['id']
            town = feat['town']
            town_parcels = feat['town_parcels']
            QgsMessageLog.logMessage('town: ' + str(town) + ' | self.town: ' + str(self.town),
                                     'BRS_GIS', level=Qgis.Info)

            if str(town_parcels) == 'NULL':
                town = town
            else:
                town = town_parcels
               
            if str(town) != str(self.town):
                pval = 'NULL'
            else:
                pval = oldPlan

            if str(jobNo) in str(pval):
                QgsMessageLog.logMessage('jobNo: ' + str(jobNo) + ' - MATCH, IGNORE!!!',
                                         'BRS_GIS', level=Qgis.Info)
                pval = ''

            if str(ppval) == str(pval):
                pval = ''

            plen = len(plans)

            QgsMessageLog.logMessage('pval: ' + str(pval), 'BRS_GIS', level=Qgis.Info)

            if str(pval) == 'NULL':
                pval = ''

                QgsMessageLog.logMessage('NULL pval: ' + str(pval), 'BRS_GIS', level=Qgis.Info)
            else:
                if pval in (plans):
                    pass
                else:
                    pval = pval
                    plans.append(pval)

            pNo += 1
            ppval = pval
            QgsMessageLog.logMessage('UAR_plans: ' + str(plans), 'BRS_GIS', level=Qgis.Info)

        feat = self.iface.activeLayer().selectedFeatures()[0]

        if len(plans) >= 1:
            plans.sort(reverse=True)
            pFinal = str(plans)
            pFinal = pFinal.strip('[')
            pFinal = pFinal.strip(']')
            pFinal = pFinal.replace('\'', '')

            if str(pFinal) == '':
                pFinal = 'N/A'

            plans = []
            #QgsMessageLog.logMessage('pFinal: ' + str(pFinal), 'BRS_GIS', level=Qgis.Info)

        else:
            plans = []
            pFinal = 'N/A'
            pass
        layerRelated.setSubsetString('id > 1')

        layerRelated.setSubsetString(u'"map_bk_lot" = \'%s\'' % (mbl))
        exp = QgsExpression(u'"map_bk_lot" = \'%s\'' % (mbl))
        request = QgsFeatureRequest(exp)
        request.setSubsetOfAttributes(['gid'], layerRelated.fields())
        request.setFlags(QgsFeatureRequest.NoGeometry)
        return pFinal
